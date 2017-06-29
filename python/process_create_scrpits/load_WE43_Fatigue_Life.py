import openpyxl
import datetime
from mcapi import create_project, get_all_templates


class WorkflowBuilderWE43FatigueLife:
    def parse_sheet(self, initiation_data_sheet, runout_data_sheet):
        initiation_data = self.parse_initiation_data(initiation_data_sheet)
        runout_data = self.parse_runout_data(runout_data_sheet)
        self.data = self.combine_data(initiation_data, runout_data)
        for sample in self.data:
            print "------ " + sample + " ------"
            tests = self.data[sample]
            for test in tests:
                print test['testing_date']
                print test['test_type']
                print '<>'
        return self

    def build_workflow(self):
        print "build_workflow"
        self.template_table = self.make_template_table()

        self.create_project_experiment()

        for sample in self.data:
            initiation_data_list = self.data[sample]['initiation']
            runout_data_list = self.data[sample]['runout']
            if len(initiation_data_list) == 1 and len(runout_data_list) > 0:
                self.create_workflow(sample, initiation_data_list[0],runout_data_list)

        print self.project.name

    def create_project_experiment(self):
        time_stamp = '%s' % datetime.datetime.now()
        project_name = "WE43 Fatigue Life: " + time_stamp
        project_description = "WE43 Fatigue Life, created from script on " + time_stamp
        self.project = create_project(project_name, project_description)

        self.experiment = self.project.create_experiment("WE43 Fatigue Life", "Fatigue measurements on WE4s samples")

    def create_workflow(self, sample_name, initiation_data, runout_data_list):
        sample = self.create_sample(sample_name, initiation_data)
        if initiation_data['surface_condition'] == 'Electropolished':
            sample = self.create_electropolishing(sample_name,sample)
        self.create_initial_measure(sample_name,sample,initiation_data)

        for data in runout_data_list:
            self.create_runout_measure(sample_name,sample,data)

    def create_sample(self, sample_name, initiation_data):
        process_name = 'Create ' + sample_name
        template_id = self.template_id_with(self.template_table, 'Create Samples')
        process = self.experiment.create_process_from_template(template_id)
        process.rename(process_name)
        annotation = "Condition: " + initiation_data['material_condition']
        process = process.add_to_notes(annotation)
        samples = process.create_samples(sample_names=[sample_name])
        return samples[0]

    def create_electropolishing(self,sample_name,sample):
        process_name = 'Electropolish ' + sample_name
        template_id = self.template_id_with(self.template_table, 'Electropolishing')
        process = self.experiment.create_process_from_template(template_id)
        process.rename(process_name)
        process.add_input_samples_to_process([sample])
        process = self.project.get_process_by_id(process.id)
        return process.output_samples[0]

    def create_initial_measure(self, sample_name, sample, initiation_data):
        process_name = "Fatigue Life Test " + sample_name
        template_id = self.template_id_with(self.template_table, 'Ultrasonic Fatigue')
        process = self.experiment.create_process_from_template(template_id)
        process.rename(process_name)
        process.add_input_samples_to_process([sample])

    def create_runout_measure(self, sample_name, sample, runout_data):
        process_name = 'Runout Fatigue Test ' + sample_name \
                       + "; Stress: " + str(runout_data['test_parameters']['stress'])
        template_id = self.template_id_with(self.template_table, 'Ultrasonic Fatigue')
        process = self.experiment.create_process_from_template(template_id)
        process.rename(process_name)
        process.add_input_samples_to_process([sample])

    def parse_initiation_data(self, data_sheet):
        list = []
        for col in data_sheet.iter_cols(min_col=2):
            values = []
            for cell in col:
                values.append(cell.value)
            if values[0]:
                list.append(self.make_initiation_data_entry(values))
        return list

    def parse_runout_data(self, data_sheet):
        list = []
        for col in data_sheet.iter_cols(min_col=2):
            values = []
            for cell in col:
                values.append(cell.value)
            if values[0]:
                list.append(self.make_runout_data_entry(values))
        return list

    def combine_data(self, initiation_data, runout_data):
        table = {}
        for element in runout_data:
            element['test_type'] = "runout"
            sample = element['sample']
            if not sample in table:
                table[sample] = []
            table[sample].append(element)
        for element in initiation_data:
            element['test_type'] = "crack initiation"
            sample = element['sample']
            if not sample in table:
                table[sample] = []
            table[sample].append(element)
        for sample in table:
            list = table[sample]
            list.sort(key=lambda e: e['testing_date'])
        return table

    def make_initiation_data_entry(self, values):
        return {
            "sample": values[0],
            "testing_date": values[1],
            "material_condition": values[2],
            "surface_condition": values[3],
            "test_parameters": {
                "stress": values[6],
                "lifetime": values[7],
                "environment": values[8],
                "step_sample": values[9],
                "control_unit": values[10],
            },
            "site_parameters": {
                "location_on_gage": values[13],
                "type": values[14],
                "location": values[15],
                "fracture_surface_character": values[16],
                "load_direction_inclination": values[17],
                "surface_inclination": values[18]
            },
            "grain_interior": {
                "schmid_factor": values[21],
                "dimension": values[22],
                "depth": values[23]
            },
            "grain_boundary": {
                "delta": values[26],
                "character": values[27]
            },
            "plotting": values[30],
            "notes": values[32]
        }

    def make_runout_data_entry(self, values):
        return {
            "sample": values[0],
            "testing_date": values[1],
            "material_condition": values[2],
            "surface_condition": values[3],
            "test_parameters": {
                "stress": values[6],
                "cycle_count": values[7],
                "environment": values[8],
                "step_sample": values[9],
                "control_unit": values[10],
            },
            "plotting": values[13],
            "notes": values[15]
        }

    def make_template_table(self):
        template_list = get_all_templates()
        table = {}
        for template in template_list:
            table[template.id] = template
        return table

    def template_id_with(self, table, match):
        found_id = None
        for key in table:
            if match in key:
                found_id = key
        return found_id


def main():
    path_and_name = 'input_data/FromJake/Fatigue Life/WE43 Fatigue Life.xlsx'
    wb = openpyxl.load_workbook(filename=path_and_name)
    ws1 = wb['Initiation Data']
    ws2 = wb['Runout Data']
    WorkflowBuilderWE43FatigueLife().parse_sheet(ws1, ws2) # \
        # .build_workflow()


if __name__ == '__main__':
    main()
