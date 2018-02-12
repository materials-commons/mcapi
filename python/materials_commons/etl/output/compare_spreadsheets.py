import argparse
import datetime
import sys
import os
from pathlib import Path

import openpyxl
from dateutil import parser as date_parser
from materials_commons.api import get_all_projects
from materials_commons.etl.common.worksheet_data import read_entire_sheet
from materials_commons.etl.input.metadata import Metadata
from .meta_data_verify import MetadataVerification


class Compare:

    def __init__(self):
        self.project = None
        self.experiment = None
        self.metadata = Metadata()

    def compare(self, project_name, experiment_name, input_file_path, output_file_path):

        ok = self.set_up_project_experiment_metadata(project_name, experiment_name)
        if not ok:
            return

        print('Input --', input_file_path)
        wb1 = openpyxl.load_workbook(filename=input_file_path)
        sheets = wb1.sheetnames
        print("Input -- Selecting: ", sheets[0], "(from", sheets, ")")
        ws1 = wb1[sheets[0]]
        data1 = read_entire_sheet(ws1)
        data1 = self.check_for_end_tag(data1)
        print("Input data size:", len(data1), len(data1[0]))

        print('Output --', output_file_path)
        wb2 = openpyxl.load_workbook(filename=output_file_path)
        sheets = wb2.sheetnames
        print("Output -- Selecting: ", sheets[0], "(from", sheets, ")")
        ws2 = wb2[sheets[0]]
        data2 = read_entire_sheet(ws2)
        data2 = self.check_for_end_tag(data2)
        print("Output data size:", len(data2), len(data2[0]))

        metadata = self.metadata
        print('----')
        if self.compare_data_shape(data1, data2):
            self.compare_headers(metadata, data1, data2)
            self.compare_first_col(metadata, data1, data2)
            self.compare_data_area(metadata, data1, data2)

    def set_up_project_experiment_metadata(self, project_name, experiment_name):
        project_list = get_all_projects()
        for proj in project_list:
            if proj.name == project_name:
                self.project = proj
        if not self.project:
            print("Can not find project with name = " + str(project_name) + ". Quiting.")
            return False
        experiment_list = self.project.get_all_experiments()
        found = []
        for exp in experiment_list:
            if exp.name == experiment_name:
                found.append(exp)
        if not found:
            print("Can not find Experiment with name = " + str(experiment_name) + ". Quiting.")
            return False
        if len(found) > 1:
            print("Found more the one Experiment with name = " + str(experiment_name) + ";")
            print("Rename experiment so that '" + str(experiment_name) + "' is unique.")
            print("Quiting.")
            return False
        self.experiment = found[0]
        ok = self.metadata.read(self.experiment.id)
        if not ok:
            print("There was no ETL metadata for the experiment '" + str(experiment_name) + "';")
            print("This experiment does not appear to have been created using ETL input.")
            print("Quiting.")
            return False
        metadata = MetadataVerification().verify(self.metadata)
        if not metadata:
            return False
        self.metadata = metadata
        return True

    @staticmethod
    def compare_data_shape(data1, data2):
        mismatch = False
        if len(data1) == 0 or len(data1) == 0:
            if len(data1) == 0 and len(data1):
                print("No data in either spreadsheet (zero length)")
            if len(data1) == 0:
                print("No data in input spreadsheet (zero length)")
            else:
                print("No data in output spreadsheet (zero length)")
            return False
        else:
            if not len(data1) == len(data2):
                print("Number of data rows differ: ", len(data1), len(data2))
                mismatch = True
            if not len(data1[0]) == len(data2[0]):
                print("Number of data cols differ: ", len(data1[0]), len(data2[0]))
                mismatch = True
            if not mismatch:
                print("Data number of rows and cols match")
        return not mismatch

    @staticmethod
    def compare_headers(metadata, data1, data2):
        len1 = len(data1)
        len2 = len(data2)
        if len1 < metadata.header_row_end:
            print("Missing header data (length), input")
        if len2 < metadata.header_row_end:
            print("Missing header data (length), output")
        if len1 >= metadata.header_row_end and len2 >= metadata.header_row_end:
            row_length_check = True
            for row in range(0, metadata.header_row_end):
                check1 = len(data1[row]) >= metadata.data_col_end
                check2 = len(data2[row]) >= metadata.data_col_end
                if not check1:
                    print("Header row shorter then expected, input, row " + str(row))
                if not check2:
                    print("Header row shorter then expected, output, row " + str(row))
                row_length_check = row_length_check and check1 and check2
            identical = True
            if row_length_check:
                for row in range(0, metadata.header_row_end):
                    row_data1 = data1[row]
                    row_data2 = data2[row]
                    for col in range(0, metadata.data_col_end):
                        match = (row_data1[col] == row_data2[col])
                        identical = identical and match
                        if not match:
                            print("Header mismatch at row = " + str(row) + ", col = " + str(col) + ": "
                                  + str(row_data1[col]) + ", " + str(row_data2[col]))
            if identical:
                print("Headers match")

    @staticmethod
    def compare_first_col(metadata, data1, data2):
        len1 = len(data1)
        len2 = len(data2)
        if len1 < metadata.data_row_end:
            print("Missing rows (first col matching), input " + str(len1))
        if len2 < metadata.data_row_end:
            print("Missing rows (first col matching), input " + str(len2))
        if len1 >= metadata.data_row_end and len2 >= metadata.data_row_end:
            identical = True
            for row in range(0, metadata.data_row_end):
                row_data1 = data1[row]
                row_data2 = data2[row]
                match = (row_data1[0] == row_data2[0])
                identical = identical and match
                if not match:
                    print("First col mismatch at row = " + str(row) + ": "
                          + str(row_data1[0]) + ", " + str(row_data2[0]))
            if identical:
                print("First cols match")

    @staticmethod
    def check_for_end_tag(data):
        first_row = data[0]
        index = 0
        missing_end = True
        end_col = len(first_row)
        for col in first_row:
            if str(col).startswith("END"):
                print("Found END marker at column " + str(index)
                      + ", updating data end to this location")
                end_col = index
                missing_end = False
                break
            index += 1
        if missing_end:
            return data
        # else
        update = []
        for data_row in data:
            update_row = []
            for index in range(0, end_col):
                update_row.append(data_row[index])
            update.append(update_row)
        return update

    def compare_data_area(self, metadata, data1, data2):
        len1 = min(len(data1), metadata.data_row_end)
        len2 = min(len(data2), metadata.data_row_end)

        end_row = len1
        if not (len1 == len2):
            print("Data number of rows differ: " + str(len1) + ", " + str(len2))
        if len2 < len1:
            end_row = len2

        if end_row < metadata.data_row_end:
            if len1 < metadata.data_row_end:
                print("Missing data rows, input, expected "
                      + str(metadata.data_row_end) + ", found" + str(len1))
            if len2 < metadata.data_row_end:
                print("Missing data rows, output, expected "
                      + str(metadata.data_row_end) + ", found" + str(len2))
        types1 = data1[metadata.header_row_end - 2]
        types2 = data2[metadata.header_row_end - 2]

        types = [""]  # element at index zero is ignored

        for col in range(1, metadata.data_col_end):
            type1 = types1[col]
            type2 = types2[col]
            data_type = type1
            if not type1 == type2:
                print("Data type mismatch, col " + str(col) + ":", type1, type2)
                if self.type_expect_data(type1):
                    data_type = type1
                elif self.type_expect_data(type2):
                    data_type = type2
            types.append(data_type)

        identical = True
        for row in range(metadata.data_row_start, end_row):
            row_data1 = data1[row]
            row_data2 = data2[row]
            row_len1 = min(len(row_data1), metadata.data_col_end)
            row_len2 = min(len(row_data2), metadata.data_col_end)
            if not (row_len1 == row_len2):
                print("Data row " + str(row) + ", lengths differ: "
                      + str(row_len1) + ", " + str(row_len2))
            end_col = row_len1
            if end_col < metadata.data_col_end:
                if row_len1 < metadata.data_col_end:
                    print("Missing data - row " + str(row) + " is short, data1, expected "
                          + str(metadata.data_col_end) + ", found " + str(row_len1))
                    print(row_data1)
                if row_len2 < metadata.data_col_end:
                    print("Missing data - row " + str(row) + " is short, data2, expected "
                          + str(metadata.data_col_end) + ", found " + str(row_len2))
                    print(row_data2)
            if row_len2 < row_len1:
                end_col = row_len2

            for col in range(1, end_col):
                if not self.type_expect_data(types[col]):
                    continue
                probe1 = row_data1[col]
                probe2 = row_data2[col]
                if isinstance(probe1, datetime.datetime):
                    if isinstance(probe2, str):
                        probe2 = date_parser.parse(probe2)
                    match = probe1.isoformat() == probe2.isoformat()
                else:
                    match = (probe1 == probe2)
                identical = identical and match
                if not match:
                    obj_type = type(probe1)
                    print("Data mismatch at row = " + str(row) + ", col = " + str(col) + ": "
                          + str(probe1) + ", " + str(probe2) + ", " + str(obj_type))

        if identical:
            print("Data values match")

    @staticmethod
    def type_expect_data(data_type):
        return data_type == "MEAS" or data_type == "PARAM" or \
               data_type == "SAMPLES" or data_type == "FILES"


def _verify_data_dir(dir_path):
    path = Path(dir_path)
    ok = path.exists() and path.is_dir()
    return ok

if __name__ == '__main__':

    argv = sys.argv
    parser = argparse.ArgumentParser(
        description='Build a workflow from given (well formatted) Excel spreadsheet')
    parser.add_argument('proj', type=str, help="Project Name")
    parser.add_argument('exp', type=str, help="Experiment Name")
    parser.add_argument('input', type=str,
                        help='Path to input EXCEL file')
    parser.add_argument('output', type=str,
                        help='Path to output EXCEL file')
    parser.add_argument('--upload', type=str,
                        help="Path to dir for uploading files; if none, files are not compared")
    parser.add_argument('--download', type=str,
                        help="Path to dir for downloaded files; if none, files are not compared")
    parser.add_argument('--checksum', action='store_true',
                        help="In comparing upload/download files, also compare checksun; optional")

    args = parser.parse_args(argv[1:])

    args.input = os.path.abspath(args.input)
    args.output = os.path.abspath(args.output)

    print("Path to input EXCEL file: " + args.input)
    print("Path to output EXCEL file: " + args.output)

    if args.upload:
        args.upload = os.path.abspath(args.upload)
        print("Path to uploaded files: " + args.upload)
    if args.download:
        args.download = os.path.abspath(args.download)
        print("Path to download files: " + args.download)

    if (args.upload or args.download):
        if not _verify_data_dir(args.upload):
            print("Path to upload directory is not valid; ignoring.")
            args.upload = ""
        if not _verify_data_dir(args.download):
            print("Path to download directory is not valid; ignoring.")
            args.download = ""
        ok = True
        if not (args.upload or args.download):
            missing = "both upload and download"
            ok = False
        elif not args.upload:
            args.upload = ""
            missing = "upload"
            ok = False
        elif not args.download:
            args.download = ""
            missing = "upload"
            ok = False
        if not ok:
            print("To compare files, you must specify both optional arguments upload and download; missing", missing)
            print("Files compare will be skipped!")
        else:
            print("Files and directories on upload and download path will be compared")
            if (args.checksum):
                print("In addition, file checksums will be computed and compared")

    c = Compare()
    c.compare(args.proj, args.exp, args.input, args.output)
