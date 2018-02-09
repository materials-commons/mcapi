import argparse
import datetime
import os
import sys

import openpyxl

from materials_commons.api import get_all_projects
from materials_commons.etl.common.util import _normalise_property_name
from materials_commons.etl.input.metadata import Metadata
from .meta_data_verify import MetadataVerification

class ExtractExperimentSpreadsheet:
    def __init__(self, output_file_path):
        self.metadata = Metadata()
        self.output_path = output_file_path
        self.worksheet = None
        self.workbook = None
        self.project = None
        self.experiment = None
        self.data_row_list = []

    def get_project(self):
        return self.project

    def get_experiment(self):
        return self.experiment

    def set_up_project_experiment_metadata(self, project_name, experiment_name):
        project_list = get_all_projects()
        for proj in project_list:
            if proj.name == project_name:
                self.project = proj
        if not self.project:
            print("Can not find project with name = " + str(project_name) + ". Quiting.")
            return False
        experiment_list = self.project.get_all_experiments()
        found = []
        for exp in experiment_list:
            if exp.name == experiment_name:
                found.append(exp)
        if not found:
            print("Can not find Experiment with name = " + str(experiment_name) + ". Quiting.")
            return False
        if len(found) > 1:
            print("Found more the one Experiment with name = " + str(experiment_name) + ";")
            print("Rename experiment so that '" + str(experiment_name) + "' is unique.")
            print("Quiting.")
            return False
        self.experiment = found[0]
        ok = self.metadata.read(self.experiment.id)
        if not ok:
            print("There was no ETL metadata for the experiment '" +  str(experiment_name) + "';")
            print("This experiment does not appear to have been created using ETL input.")
            print("Quiting.")
            return False
        metadata = MetadataVerification().verify(self.metadata)
        if not metadata:
            return False
        self.metadata = metadata
        return True

    def build_experiment_array(self):
        self.data_row_list = []
        self.set_headers_from_metadata()
        self.set_data_from_metadata()

    def set_headers_from_metadata(self):
        print("    ... setting headers...")
        for row in self.metadata.sheet_headers:
            self.data_row_list.append(row)

    def set_data_from_metadata(self):
        self.initialize_empty_data()
        metadata = self.metadata
        self.data_row_list[metadata.data_row_start][0] = "BEGIN_DATA"
        print("    ... setting data...")
        process_record_list = metadata.process_metadata
        table = metadata.process_table
        type_list = metadata.sheet_headers[metadata.start_attribute_row]
        attribute_list = metadata.sheet_headers[metadata.start_attribute_row + 1]
        attribute_list = self.remove_units_spec_from_attributes(attribute_list)
        attribute_list = self.parse_dotted_attributes(attribute_list)
        attribute_list = self.make_normalized_attribute_names(attribute_list)
        for process_record in process_record_list:
            start_row = process_record["start_row"]
            end_row = process_record["end_row"]
            start_col = process_record["start_col"]
            end_col = process_record["end_col"]
            process = table[process_record['id']]
            self.write_first_data_row_for_process(
                start_row, start_col, end_col, type_list, attribute_list, process)
            if (end_row - start_row) > 1:
                self.copy_duplicate_rows_for_process(
                    start_row, end_row, start_col, end_col)

    def initialize_empty_data(self):
        for row_index in range(len(self.data_row_list), self.metadata.data_row_end):
            self.data_row_list.append([None] * self.metadata.data_col_end)

    def write_first_data_row_for_process(self, row, start, end, types, attributes, process):
        measurements = process.measurements
        setup_parameter_list = process.setup
        for col in range(start, end):
            value_type = types[col]
            attribute = attributes[col]
            if value_type == "MEAS":
                value = self.extract_measurement_for(attribute, measurements)
            elif value_type == "PARAM":
                value = self.extract_parameter_for(attribute, setup_parameter_list)
            elif value_type == "SAMPLES" and process.output_samples:
                value = process.output_samples[0].name
            else:
                value = None
            self.data_row_list[row][col] = value

    def extract_measurement_for(self, attribute, measurements):
        value = None
        measurement = self.find_measurement_for_attribute(attribute, measurements)
        if measurement:
            value = self.get_measurement_value_for_attribute(attribute, measurement)
        return value

    def find_measurement_for_attribute(self, attribute, measurements):
        found_measurement = None
        base = attribute
        if not isinstance(attribute, str):
            base = attribute[0]
        for m in measurements:
            if base == m.attribute:
                found_measurement = m
        return found_measurement

    def get_measurement_value_for_attribute(self, attribute, measurement):
        # print('get_measurement_value_for_attribute', attribute, measurement.value)
        if not isinstance(attribute, list):
            if not isinstance(measurement.value, list):
                return measurement.value
            else:
                return None
        return self.recursive_value_extraction(attribute[0], attribute[1:],measurement.value)

    def recursive_value_extraction(self, name, name_list, probe):
        key = self.key_for_category(name, name_list)
        # print('recursive_value_extraction', name, name_list, key, probe)
        value = None
        if isinstance(probe, dict):
            if key in probe:
                value = probe[key]
        elif len(name_list) > 0:
            part_name = name_list[0]
            for m in probe:
                if key in m and m[key] == part_name:
                    value = m['value']
                    if isinstance(value, dict):
                        value = self.recursive_value_extraction(name_list[1], name_list[2:], value)
        # print('recursive_value_extraction - value', value)
        return value

    @staticmethod
    def key_for_category(name, name_list):
        key = name
        if name == 'composition':
            key = 'element'
        if name == 'stats':
            key = name_list[0]
        return key

    @staticmethod
    def extract_parameter_for(attribute, setup_list):
        value = None
        for s in setup_list:
            for prop in s.properties:
                if attribute.startswith(prop.attribute):
                    value = prop.value
        return value

    def copy_duplicate_rows_for_process(self, start_row, end_row, start_col, end_col):
        for row in range(start_row + 1, end_row):
            for col in range(start_col, end_col):
                self.data_row_list[row][col] = self.data_row_list[start_row][col]

    def write_spreadsheet(self):
        print("Writing spreadsheet", self.output_path)
        if self.create_worksheet():
            self.write_data_to_sheet()
            self.workbook.save(filename=self.output_path)
            self.workbook.close()

    def write_data_to_sheet(self):
        for row in range(0, len(self.data_row_list)):
            data_row = self.data_row_list[row]
            for col in range(0, len(data_row)):
                data_item = data_row[col]
                self.worksheet.cell(column=col + 1, row=row + 1, value=data_item)

    def create_worksheet(self):
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws['A1'] = "PROJ: " + self.project.name
        ws['A2'] = "EXP: " + self.experiment.name
        wb.save(filename=self.output_path)
        self.worksheet = ws
        self.workbook = wb
        return wb

    @staticmethod
    def remove_units_spec_from_attributes(attributes):
        update = []
        for attr in attributes:
            if attr and '(' in attr:
                pos = attr.find('(')
                if pos > 2:
                    attr = attr[0:pos-1]
            update.append(attr)
        return update

    @staticmethod
    def parse_dotted_attributes(attributes):
        update = []
        for attr in attributes:
            if attr and '.' in attr:
                attr = attr.split('.')
            update.append(attr)
        return update

    @staticmethod
    def make_normalized_attribute_names(attributes):
        update = []
        for attr in attributes:
            if attr:
                if isinstance(attr,str):
                    attr = _normalise_property_name(attr)
                else:
                    parts = []
                    for part in attr:
                        if not parts:
                            # only normalize the leading term
                            part = _normalise_property_name(part)
                        parts.append(part)
                    attr = parts
            update.append(attr)
        return update


def main(main_args):
    builder = ExtractExperimentSpreadsheet(main_args.output)
    ok = builder.set_up_project_experiment_metadata(main_args.proj, main_args.exp)
    if ok:
        print("Writing experiment '" + builder.experiment.name
              + "' in project, '" + builder.project.name + ", to")
        print("spreadsheet at " + builder.output_path)
        builder.build_experiment_array()
        builder.write_spreadsheet()


if __name__ == '__main__':
    time_stamp = '%s' % datetime.datetime.now()
    default_output_file_path = os.path.abspath("output.xlsx")

    argv = sys.argv
    parser = argparse.ArgumentParser(
        description='Dump a project-experiment to a spreadsheet')
    parser.add_argument('proj', type=str, help="Project Name")
    parser.add_argument('exp', type=str, help="Experiment Name")
    parser.add_argument('--output', type=str, default=default_output_file_path,
                        help='Path to output file, defaults to ' + default_output_file_path)
    args = parser.parse_args(argv[1:])

    args.output = os.path.abspath(args.output)

    if not args.output.endswith(".xlsx"):
        file = args.output + ".xlsx"

    main(args)
