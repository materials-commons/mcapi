import argparse
import datetime
import sys
import tempfile
from pathlib import Path

import openpyxl

from materials_commons.api import get_all_projects
from materials_commons.etl.common.util import _normalise_property_name
from materials_commons.etl.common.worksheet_data import read_entire_sheet
from materials_commons.etl.input.metadata import Metadata
from .meta_data_verify import MetadataVerification


class Differ:
    def __init__(self):
        self.metadata = Metadata()
        self.project = None
        self.experiment = None
        self.input_data = []
        self.missing_process_list = []
        self.added_process_list = []

    def compute_deltas(self):
        metadata = self.metadata
        process_table = metadata.process_table
        data = self.input_data
        delta_list = []
        print("Computing deltas for experiment:", self.experiment.name)
        print("excel spreadsheet...")
        print("  number of rows:", len(data))
        print("  length of first row", len(data[0]))
        print("metadata...")
        print("  data rows, start and end:", metadata.data_row_start, metadata.data_row_end)
        print("  data cols, start and end:", metadata.data_col_start, metadata.data_col_end)
        print("  number of process records:", len(metadata.process_metadata))
        print("experiment...")
        print("  number of processes", len(process_table))
        delta_list += self._process_deltas()
        delta_list += self._attribute_deltas()
        delta_list += self._value_deltas()
        return delta_list

    def report_deltas(self, deltas):
        print("Reporting deltas for experiment:", self.experiment.name)
        print("Number of deltas:", len(deltas))
        for delta in deltas:
            print("  ", delta['type'], delta['data'])

    def set_up_project_experiment_metadata(self, project_name, experiment_name):
        project_list = get_all_projects()
        for proj in project_list:
            if proj.name == project_name:
                self.project = proj
        if not self.project:
            print("Can not find project with name = " + str(project_name) + ". Quiting.")
            return False
        experiment_list = self.project.get_all_experiments()
        found = []
        for exp in experiment_list:
            if exp.name == experiment_name:
                found.append(exp)
        if not found:
            print("Can not find Experiment with name = " + str(experiment_name) + ". Quiting.")
            return False
        if len(found) > 1:
            print("Found more the one Experiment with name = " + str(experiment_name) + ";")
            print("Rename experiment so that '" + str(experiment_name) + "' is unique.")
            print("Quiting.")
            return False
        self.experiment = found[0]
        ok = self.metadata.read(self.experiment.id)
        if not ok:
            print("There was no ETL metadata for the experiment '" + str(experiment_name) + "';")
            print("This experiment does not appear to have been created using ETL input.")
            print("Quiting.")
            return False
        verify = MetadataVerification()
        metadata = verify.verify(self.metadata)  # Adds metadata.process_table !
        if not metadata:
            if verify.failure == "Project":
                print("Failed to find project for compare. Quiting")
                return False
            if verify.failure == "Experiment":
                print("Failed to find experiment for compare. Quiting")
                return False
            print("Differences detected by metadata verification")
            metadata = verify.metadata
            if verify.missing_process_list:
                self.missing_process_list = verify.missing_process_list
            if verify.added_process_list:
                self.added_process_list = verify.added_process_list
        self.metadata = metadata
        return True

    def set_up_input_data(self):
        file_id = self.metadata.input_excel_file_id
        if not file_id:
            print("Experiment metadata does not contain id of excel file")
            return False
        with tempfile.TemporaryDirectory() as tmpdirname:
            directory = self.project.add_directory("/Input Excel Spreadsheets")
            file = None
            for child in directory.get_children():
                if child.id == file_id:
                    file = child
                    break
            if not file:
                print("Can not locate file for excel input")
                return False
            else:
                local_file_path = str(Path(tmpdirname, file.name))
                file.download_file_content(local_file_path)
                wb = openpyxl.load_workbook(filename=local_file_path)
                sheet_name = wb.sheetnames[0]
                ws = wb[sheet_name]
                print("In input Excel file, using sheet", sheet_name, "from sheets", wb.sheetnames)
                self.input_data = read_entire_sheet(ws)
                wb.close()
        return True

    def _process_deltas(self):
        deltas = []
        if self.missing_process_list:
            for process_id in self.missing_process_list:
                deltas.append({
                    "type": "missing_process",
                    "data": {
                        "process_id": process_id
                    }
                })
        if self.added_process_list:
            for process_id in self.added_process_list:
                deltas.append({
                    "type": "added_process",
                    "data": {
                        "process_id": process_id
                    }
                })
        return deltas

    def _attribute_deltas(self):
        deltas = []
        for process_id in self.metadata.process_table:
            if process_id in self.missing_process_list:
                continue
            metadata_attribute_table = self._get_metadata_attributes(process_id)
            metadata_attribute_list = list(metadata_attribute_table.keys())
            process_attribute_and_type_list = self._get_process_attributes_and_type(process_id)
            process_attribute_type_table = {}
            for entry in process_attribute_and_type_list:
                process_attribute_type_table[entry['attribute']] = entry['type']
            process_attribute_list = list(process_attribute_type_table.keys())
            for attribute in metadata_attribute_list:
                if attribute not in process_attribute_list:
                    deltas.append({
                        "type": "missing_attribute",
                        "data": {
                            'process_id': process_id,
                            'attribute': attribute,
                        }
                    })
            print("process_attribute_list", process_attribute_list)
            print("metadata_attribute_list", metadata_attribute_list)
            for attribute in process_attribute_list:
                if attribute not in metadata_attribute_list:
                    deltas.append({
                        "type": "added_attribute",
                        "data": {
                            'process_id': process_id,
                            'attribute': attribute,
                            'type': process_attribute_type_table[attribute]
                        }
                    })
        return deltas

    def _value_deltas(self):
        deltas = []
        for process_id in self.metadata.process_table:
            if process_id in self.missing_process_list:
                continue
            process_record = self._get_metadata_process_record(process_id)
            types_row = self.input_data[self.metadata.start_attribute_row]
            if process_record:
                process = self.metadata.process_table[process_id]
                metadata_attribute_table = self._get_metadata_attributes(process_id)
                metadata_attribute_list = list(metadata_attribute_table.keys())
                attributes_row = self._remove_units_spec_from_attributes(
                    self.input_data[self.metadata.start_attribute_row + 1])
                attributes_row = self._make_normalized_attribute_names(attributes_row)
                start_col = process_record['start_col']
                end_col = process_record['end_col']
                measurements = self._select_best_measures(process.measurements)
                setup_parameter_list = process.setup
                for col in range(start_col, end_col):
                    value_type = types_row[col]
                    attribute = attributes_row[col]
                    if value_type not in ["MEAS", "PARAM", "SAMPLES"]:
                        continue
                    if attribute in metadata_attribute_list:
                        process_value = None
                        if value_type == "MEAS":
                            process_value = self.extract_measurement_for(attribute, measurements)
                        elif value_type == "PARAM":
                            process_value = self.extract_parameter_for(attribute, setup_parameter_list)
                        elif value_type == "SAMPLES" and process.output_samples:
                            process_value = process.output_samples[0].name
                        start_row = process_record['start_row']
                        end_row = process_record['end_row']
                        for row in range(start_row, end_row):
                            input_value = self.input_data[row][col]
                            if process_value == 'None':  # how did that happen;
                                #  TODO: track down where process_value is set to 'None'
                                process_value = None
                            if process_value == '':
                                process_value = None # another one!
                            if process_value != input_value:
                                deltas.append({
                                    "type": "changed_value",
                                    "data": {
                                        'process_id': process_id,
                                        'attribute': attribute,
                                        'type': value_type,
                                        'old_value': input_value,
                                        'new_value': process_value,
                                        'location': {'row': row, 'col': col}
                                    }
                                })
        return deltas

    @staticmethod
    def extract_parameter_for(attribute, setup_list):
        value = None
        for s in setup_list:
            for prop in s.properties:
                if attribute.startswith(prop.attribute):
                    value = prop.value
        return value

    def extract_measurement_for(self, attribute, measurements):
        value = None
        measurement = self.find_measurement_for_attribute(attribute, measurements)
        if measurement:
            value = self.get_measurement_value_for_attribute(attribute, measurement)
        return value

    @staticmethod
    def find_measurement_for_attribute(attribute, measurements):
        found_measurement = None
        base = attribute
        if not isinstance(attribute, str):
            base = attribute[0]
        for m in measurements:
            if base == m.attribute:
                found_measurement = m
        return found_measurement

    def get_measurement_value_for_attribute(self, attribute, measurement):
        # print('get_measurement_value_for_attribute', attribute, measurement.value)
        if not isinstance(attribute, list):
            if not isinstance(measurement.value, list):
                return measurement.value
            else:
                return None
        return self.recursive_value_extraction(attribute[0], attribute[1:], measurement.value)

    def recursive_value_extraction(self, name, name_list, probe):
        key = self.key_for_category(name, name_list)
        # print('recursive_value_extraction', name, name_list, key, probe)
        value = None
        if isinstance(probe, dict):
            if key in probe:
                value = probe[key]
        elif len(name_list) > 0:
            part_name = name_list[0]
            for m in probe:
                if key in m and m[key] == part_name:
                    value = m['value']
                    if isinstance(value, dict):
                        value = self.recursive_value_extraction(name_list[1], name_list[2:], value)
        # print('recursive_value_extraction - value', value)
        return value

    @staticmethod
    def key_for_category(name, name_list):
        key = name
        if name == 'composition':
            key = 'element'
        if name == 'stats':
            key = name_list[0]
        return key

    def _get_metadata_attributes(self, process_id):
        attribute_table = {}
        process_record = self._get_metadata_process_record(process_id)
        types_row = self.input_data[self.metadata.start_attribute_row]
        attributes_row = self._remove_units_spec_from_attributes(
            self.input_data[self.metadata.start_attribute_row + 1])
        attributes_row = self._make_normalized_attribute_names(attributes_row)
        if process_record:
            start_col = process_record['start_col']
            end_col = process_record['end_col']
            row = process_record['start_row']
            for col in range(start_col, end_col):
                if self._is_attribute(types_row[col]):
                    attribute_table[attributes_row[col]] = {"with_data": self._is_data(row, col)}
        # print("  for process id", process_id, "metadata attributes are", attribute_list)
        return attribute_table

    def _get_process_attributes_and_type(self, process_id):
        attribute_and_type_list = []
        process = self.metadata.process_table[process_id]
        for s in process.setup:
            for prop in s.properties:
                if (prop.value is not None) and (str(prop.value).strip() != ""):
                    attribute_and_type_list.append({'attribute': prop.attribute, 'type': 'PARAM'})
        for m in process.measurements:
            attribute_and_type_list.append({'attribute': m.attribute, 'type': 'MEAS'})
        return attribute_and_type_list

    def _get_metadata_process_record(self, process_id):
        found = None
        for record in self.metadata.process_metadata:
            if process_id == record['id']:
                found = record
                break
        return found

    def _is_data(self, row, col):
        value = self.input_data[row][col]
        results = not (value is None or value == '')
        return results

    @staticmethod
    def _select_best_measures(measurements):
        selected_measurements = {}
        for m in measurements:
            if m.attribute in selected_measurements:
                sm = selected_measurements[m.attribute]
                if sm.is_best_measure:
                    continue
            else:
                selected_measurements[m.attribute] = m
        results = []
        for key in selected_measurements:
            results.append(selected_measurements[key])
        return results

    @staticmethod
    def _is_attribute(attribute_type):
        return attribute_type in ['PARAM', 'MEAS']

    @staticmethod
    def _remove_units_spec_from_attributes(attributes):
        update = []
        for attr in attributes:
            if attr and '(' in attr:
                pos = attr.find('(')
                if pos > 2:
                    attr = attr[0:pos - 1]
            update.append(attr)
        return update

    @staticmethod
    def _make_normalized_attribute_names(attributes):
        update = []
        for attr in attributes:
            if attr:
                if isinstance(attr, str):
                    attr = _normalise_property_name(attr)
                else:
                    parts = []
                    for part in attr:
                        if not parts:
                            # only normalize the leading term
                            part = _normalise_property_name(part)
                        parts.append(part)
                    attr = parts
            update.append(attr)
        return update


def main(project_name, experiment_name):
    differ = Differ()
    ok = differ.set_up_project_experiment_metadata(project_name, experiment_name)
    if not ok:
        print("Invalid configuration of metadata or experiment/metadata mismatch. Quiting")
        exit(-1)
    ok = differ.set_up_input_data()
    if not ok:
        print("Could not locate or read input spreadsheet from experiment")
        exit(-1)
    deltas = differ.compute_deltas()
    if not deltas:
        print("No differences were detected. Done")
    else:
        differ.report_deltas(deltas)


if __name__ == '__main__':
    time_stamp = '%s' % datetime.datetime.now()

    argv = sys.argv
    parser = argparse.ArgumentParser(
        description='Computer differences in web site and excel spreadsheet data for experiment')
    parser.add_argument('proj', type=str, help="Project Name")
    parser.add_argument('exp', type=str, help="Experiment Name")
    args = parser.parse_args(argv[1:])

    print(
        "Computer differences in web site and excel spreadsheet data for experiment '"
        + args.exp + "' of project '" + args.proj)

    main(args.proj, args.exp)
