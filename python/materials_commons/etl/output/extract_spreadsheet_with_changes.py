import argparse
import datetime
import json
import os
import sys
from pathlib import Path
from pprint import pprint

import openpyxl

from materials_commons.api import File as MC_File
from materials_commons.api import get_all_projects
from materials_commons.etl.common.process_file_util import make_project_file_id_path_table
from materials_commons.etl.common.util import _normalise_property_name
from materials_commons.etl.input.metadata import Metadata
from .meta_data_verify import MetadataVerification


class ExtractExperimentSpreadsheetWithChanges:
    def __init__(self, output_file_path):
        self.metadata = Metadata()
        self.output_path = output_file_path
        self.worksheet = None
        self.workbook = None
        self.project = None
        self.experiment = None
        self.delta_list = None
        self.data_built = False
        self.data_row_list = []

    def get_project(self):
        return self.project

    def get_experiment(self):
        return self.experiment

    def set_up_project_experiment_metadata(self, project_name, experiment_name):
        project_list = get_all_projects()
        for proj in project_list:
            if proj.name == project_name:
                self.project = proj
        if not self.project:
            print("Can not find project with name = " + str(project_name) + ". Quiting.")
            return False
        experiment_list = self.project.get_all_experiments()
        found = []
        for exp in experiment_list:
            if exp.name == experiment_name:
                found.append(exp)
        if not found:
            print("Can not find Experiment with name = " + str(experiment_name) + ". Quiting.")
            return False
        if len(found) > 1:
            print("Found more the one Experiment with name = " + str(experiment_name) + ";")
            print("Rename experiment so that '" + str(experiment_name) + "' is unique.")
            print("Quiting.")
            return False
        self.experiment = found[0]
        ok = self.metadata.read(self.experiment.id)
        if not ok:
            print("There was no ETL metadata for the experiment '" + str(experiment_name) + "';")
            print("This experiment does not appear to have been created using ETL input.")
            print("Quiting.")
            return False
        verifier = MetadataVerification(self.metadata)
        verifier.add_process_table_to_metadata(self.experiment)
        if self.delta_list:
            metadata = verifier.verify_with_delta_list(self.delta_list)
        else:
            metadata = verifier.verify()
        if not metadata:
            print("Metadata verification failed.")
            return False
        self.metadata = metadata
        return True

    def register_deltas(self, deltas):
        self.delta_list = deltas

    def build_experiment_array(self):
        self.data_row_list = []
        self.set_headers_from_metadata()
        self.set_data_from_metadata()

    def apply_deltas(self):
        if not self.delta_list:
            print("Attempted to apply deltas, but no delta list, call register_deltas() first")
            return
        if not self.data_built:
            print("Attempt to apply delta list, but data not build, call build_experiment_array() first")
            return
        # NOTE: missing processes already handled
        # NOTE: missing attributes can be, and are, ignored
        self.apply_changed_values()
        self.apply_added_attributes()
        self.apply_added_processes()

    def set_headers_from_metadata(self):
        print("    ... setting headers...")
        for row in self.metadata.sheet_headers:
            self.data_row_list.append(row)

    def set_data_from_metadata(self):
        self.initialize_empty_data()
        metadata = self.metadata
        self.data_row_list[metadata.data_row_start][0] = "BEGIN_DATA"
        print("    ... setting data...")
        process_record_list = metadata.process_metadata
        table = metadata.process_table  # Note: added by metadata verify
        type_list = metadata.sheet_headers[metadata.start_attribute_row]
        attribute_list = metadata.sheet_headers[metadata.start_attribute_row + 1]
        attribute_list = self.remove_units_spec_from_attributes(attribute_list)
        attribute_list = self.parse_dotted_attributes(attribute_list)
        attribute_list = self.make_normalized_attribute_names(attribute_list)
        for process_record in process_record_list:
            start_row = process_record["start_row"]
            end_row = process_record["end_row"]
            start_col = process_record["start_col"]
            end_col = process_record["end_col"]
            process_id = process_record['id']
            if self.check_delta_list_for_missing_process(process_id):
                print("Accepting missing process (in delta list)", process_id)
                continue
            process = table[process_record['id']]
            self.write_first_data_row_for_process(
                start_row, start_col, end_col, type_list, attribute_list, process)
            self.set_file_entry_for_process(
                start_row, start_col, end_col, type_list, process_record['files']
            )
            if (end_row - start_row) > 1:
                self.copy_duplicate_rows_for_process(
                    start_row, end_row, start_col, end_col)
        self.data_built = True

    def initialize_empty_data(self):
        for row_index in range(len(self.data_row_list), self.metadata.data_row_end):
            self.data_row_list.append([None] * self.metadata.data_col_end)

    def write_first_data_row_for_process(self, row, start, end, types, attributes, process):
        measurements = process.measurements
        setup_parameter_list = process.setup
        for col in range(start, end):
            value_type = types[col]
            attribute = attributes[col]
            if value_type == "MEAS":
                value = self.extract_measurement_for(attribute, measurements)
            elif value_type == "PARAM":
                value = self.extract_parameter_for(attribute, setup_parameter_list)
            elif value_type == "SAMPLES" and process.output_samples:
                value = process.output_samples[0].name
            else:
                value = None
            self.data_row_list[row][col] = value

    def apply_changed_values(self):
        # Note, this must be applied before other deltas, as other update may shift data locations!
        for item in self.delta_list:
            if not item['type'] == 'changed_value':
                continue
            row = item['data']['location']['row']
            col = item['data']['location']['col']
            value = item['data']['new_value']
            self.data_row_list[row][col] = value

    def apply_added_attributes(self):
        for item in self.delta_list:
            if not item['type'] == 'added_attribute':
                continue
            process_id = item['data']['process_id']
            attribute = item['data']['attribute']
            attribute_type = item['data']['type']
            process = self.metadata.process_table[process_id]
            measurements = process.measurements
            setup_parameter_list = process.setup
            if attribute_type == "MEAS":
                value = self.extract_measurement_for(attribute, measurements)
            elif attribute_type == "PARAM":
                value = self.extract_parameter_for(attribute, setup_parameter_list)
            else:
                value = None
            print("apply_added_attributes", process_id, attribute_type, attribute, value)
            self._insert_new_attribute_in_data(process_id, attribute_type, attribute, value)

    def apply_added_processes(self):
        for item in self.delta_list:
            if not item['type'] == 'added_process':
                continue
            process_id = item['data']['process_id']
            process = self.metadata.process_table[process_id]
            proc_name = process.template_name.replace("Template","").strip()
            print(process.input_data)
            attribute_list = []
            measurements = process.measurements
            setup_parameter_list = process.setup
            for m in measurements:
                attribute = m.attribute
                value = self.get_measurement_value_for_attribute(attribute, m)
                attribute_list.append({'type': 'MEAS', 'attribute': attribute, 'value': value})
            for p in setup_parameter_list:
                for prop in p.properties:
                    attribute = prop.attribute
                    value = prop.value
                    if (value is None) or (value == ""):
                        continue
                    attribute_list.append({'type': 'PARAM', 'attribute': attribute, 'value': value})
            print("apply_added_processes", process.id, process.name, attribute_list)
            parent = None
            sample = process.input_samples[0]
            for probe_id in self.metadata.process_table:
                probe = self.metadata.process_table[probe_id]
                if probe.output_samples:
                    probe_sample = probe.output_samples[0]
                    if probe_sample.id == sample.id and probe_sample.property_set_id == sample.property_set_id:
                        parent = probe
            parent_id = None
            if parent:
                parent_id = parent.id
            print("parent", parent_id)
            new_col = self.metadata.data_col_start
            row = self.metadata.data_row_start
            if parent_id:
                for probe in self.metadata.process_metadata:
                    if probe['id'] == parent_id:
                        new_col = probe['end_col']
                        row = probe['start_row']
            print("location", row, new_col)
            for process_entry in self.metadata.process_metadata:
                probe_start_col = process_entry['start_col']
                if probe_start_col >= new_col:
                    process_entry['end_col'] = process_entry['end_col'] + 1
                    if probe_start_col > new_col:
                        process_entry['start_col'] = process_entry['start_col'] + 1
            new_process_entry = {
                'id': process.id,
                'name': process.name,
                'template': process.template_name,
                'files': "",
                'start_row': row,
                'end_row': row+1,
                'start_col': new_col,
                'end_col': new_col+1
            }
            placed = False,
            new_process_metadata = []
            for process_entry in self.metadata.process_metadata:
                probe_start_col = process_entry['start_col']
                probe_start_row = process_entry['start_row']
                if (not placed) and (new_col >= probe_start_col) and (row > probe_start_row):
                    new_process_metadata.append(new_process_entry)
                    placed = True
                new_process_metadata.append(process_entry)
            if not placed:
                new_process_metadata.append(new_process_entry)
            self.metadata.process_metadata = new_process_metadata

            new_sheet_headers = []
            for header in self.metadata.sheet_headers:
                pre = header[:new_col]
                post = header[new_col:]
                new_header = pre + ['null'] + post
                new_sheet_headers.append(new_header)
            new_sheet_headers[0][new_col] = proc_name
            new_sheet_headers[self.metadata.start_attribute_row][new_col] = "IGNORE"
            self.metadata.sheet_headers = new_sheet_headers
            print(self.metadata.sheet_headers)
            new_data_row_list = []
            for data_row in self.data_row_list:
                pre = data_row[:new_col]
                post = data_row[new_col:]
                new_data_row = pre + [''] + post
                new_data_row_list.append(new_data_row)
            self.data_row_list = new_data_row_list
            self.data_row_list[0][new_col] = proc_name
            self.data_row_list[self.metadata.start_attribute_row][new_col] = "IGNORE"
            self.data_row_list[row][new_col] = "Mark"
            print(self.data_row_list)

    def set_file_entry_for_process(self, row, start, end, types, entry):
        for col in range(start, end):
            value_type = types[col]
            if value_type == "FILES":
                self.data_row_list[row][col] = entry
                break

    def extract_measurement_for(self, attribute, measurements):
        value = None
        measurement = self.find_measurement_for_attribute(attribute, measurements)
        if measurement:
            value = self.get_measurement_value_for_attribute(attribute, measurement)
        return value

    @staticmethod
    def find_measurement_for_attribute(attribute, measurements):
        found_measurement = None
        base = attribute
        if not isinstance(attribute, str):
            base = attribute[0]
        for m in measurements:
            if base == m.attribute:
                found_measurement = m
        return found_measurement

    def get_measurement_value_for_attribute(self, attribute, measurement):
        # print('get_measurement_value_for_attribute', attribute, measurement.value)
        if not isinstance(attribute, list):
            if not isinstance(measurement.value, list):
                return measurement.value
            else:
                return None
        return self.recursive_value_extraction(attribute[0], attribute[1:], measurement.value)

    def recursive_value_extraction(self, name, name_list, probe):
        key = self.key_for_category(name, name_list)
        # print('recursive_value_extraction', name, name_list, key, probe)
        value = None
        if isinstance(probe, dict):
            if key in probe:
                value = probe[key]
        elif len(name_list) > 0:
            part_name = name_list[0]
            for m in probe:
                if key in m and m[key] == part_name:
                    value = m['value']
                    if isinstance(value, dict):
                        value = self.recursive_value_extraction(name_list[1], name_list[2:], value)
        # print('recursive_value_extraction - value', value)
        return value

    def download_process_files(self, download_dir_path):
        metadata = self.metadata
        project = self.get_project()
        top_dir = project.get_top_directory()
        top_directory_name = top_dir.name
        file_id_path_table = make_project_file_id_path_table(project)
        path_table = {}
        for key in file_id_path_table:
            item = file_id_path_table[key]
            # path = item['path'] # not used
            path_table[item['path']] = item
        project.local_path = download_dir_path
        process_record_list = metadata.process_metadata
        type_list = metadata.sheet_headers[metadata.start_attribute_row]
        for process_record in process_record_list:
            start_row = process_record["start_row"]
            start_col = process_record["start_col"]
            end_col = process_record["end_col"]
            files_entry = None
            for col in range(start_col, end_col):
                if type_list[col] == "FILES":
                    files_entry = self.data_row_list[start_row][col]
            if files_entry:
                file_or_dir_path_list = files_entry.split(",")
                for entry in file_or_dir_path_list:
                    entry = entry.strip()
                    path = Path(self.project.local_path) / entry
                    local_path = str(path)
                    remote_path = "/" + str(Path(top_directory_name) / entry)
                    print("  : ", remote_path, "-->", local_path)
                    if remote_path not in path_table:
                        print("  :     file from spreadsheet not in project, skipping", entry)
                        continue
                    record = path_table[str(remote_path)]
                    if record['is_file']:
                        self.download_local_file_content(record['file'], path)
                    else:
                        self.download_local_dir_content(record['dir'], path)

    @staticmethod
    def download_local_file_content(target_file, path):
        if path.exists():
            print("  :     skipping duplicate: ", path)
        else:
            parent_dir = Path(*list(path.parts)[:-1])
            os.makedirs(str(parent_dir), exist_ok=True)
            target_file.download_file_content(str(path))

    def download_local_dir_content(self, target_dir, path):
        #        print("download dir", dir.name, path)
        os.makedirs(path, exist_ok=True)
        for child in target_dir.get_children():
            child_path = Path(path, child.name)
            #            print("child_path", str(child_path))
            if type(child) == MC_File:
                self.download_local_file_content(child, child_path)
            else:
                self.download_local_dir_content(child, child_path)

    def get_deltas_from_file(self, file_path):
        deltas = None
        if Path(file_path).exists():
            delta_file = open(file_path)
            deltas = json.load(delta_file)
            delta_file.close()
        if not deltas:
            return False
        pprint(deltas)
        self.register_deltas(deltas)
        return True

    @staticmethod
    def key_for_category(name, name_list):
        key = name
        if name == 'composition':
            key = 'element'
        if name == 'stats':
            key = name_list[0]
        return key

    @staticmethod
    def extract_parameter_for(attribute, setup_list):
        value = None
        for s in setup_list:
            for prop in s.properties:
                if attribute.startswith(prop.attribute):
                    value = prop.value
        return value

    def copy_duplicate_rows_for_process(self, start_row, end_row, start_col, end_col):
        for row in range(start_row + 1, end_row):
            for col in range(start_col, end_col):
                self.data_row_list[row][col] = self.data_row_list[start_row][col]

    def write_spreadsheet(self):
        print("Writing spreadsheet", self.output_path)
        if self.create_worksheet():
            self.write_data_to_sheet()
            self.workbook.save(filename=self.output_path)
            self.workbook.close()

    def write_data_to_sheet(self):
        for row in range(0, len(self.data_row_list)):
            data_row = self.data_row_list[row]
            for col in range(0, len(data_row)):
                data_item = data_row[col]
                self.worksheet.cell(column=col + 1, row=row + 1, value=data_item)

    def create_worksheet(self):
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]
        ws['A1'] = "PROJ: " + self.project.name
        ws['A2'] = "EXP: " + self.experiment.name
        wb.save(filename=self.output_path)
        self.worksheet = ws
        self.workbook = wb
        return wb

    @staticmethod
    def remove_units_spec_from_attributes(attributes):
        update = []
        for attr in attributes:
            if attr and '(' in attr:
                pos = attr.find('(')
                if pos > 2:
                    attr = attr[0:pos - 1]
            update.append(attr)
        return update

    @staticmethod
    def parse_dotted_attributes(attributes):
        update = []
        for attr in attributes:
            if attr and '.' in attr:
                attr = attr.split('.')
            update.append(attr)
        return update

    @staticmethod
    def make_normalized_attribute_names(attributes):
        update = []
        for attr in attributes:
            if attr:
                if isinstance(attr, str):
                    attr = _normalise_property_name(attr)
                else:
                    parts = []
                    for part in attr:
                        if not parts:
                            # only normalize the leading term
                            part = _normalise_property_name(part)
                        parts.append(part)
                    attr = parts
            update.append(attr)
        return update

    @staticmethod
    def make_process_filenames(process):
        files = process.get_all_files()
        if not files:
            return ""
        names = []
        for f in files:
            names.append(f.name)
        return ", ".join(names)

    def check_delta_list_for_missing_process(self, process_id):
        for item in self.delta_list:
            if item['type'] == 'missing_process' and process_id == item['data']['process_id']:
                return True
        return False

    def check_delta_list_for_added_process(self, process_id):
        for item in self.delta_list:
            if item['type'] == 'added_process' and process_id == item['data']['process_id']:
                return True
        return False

    def _insert_new_attribute_in_data(self, process_id, attribute_type, attribute, value):
        self.metadata.data_col_end = self.metadata.data_col_end + 1
        process_in_metadata = None
        for process_entry in self.metadata.process_metadata:
            if process_entry['id'] == process_id:
                process_in_metadata = process_entry
        start_col = process_in_metadata['start_col']
        end_col = process_in_metadata['end_col']
        start_row = process_in_metadata['start_row']
        end_row = process_in_metadata['end_row']
        for process_entry in self.metadata.process_metadata:
            probe_start_col = process_entry['start_col']
            if probe_start_col >= start_col:
                process_entry['end_col'] = process_entry['end_col'] + 1
                if probe_start_col > start_col:
                    process_entry['start_col'] = process_entry['start_col'] + 1
        new_sheet_headers = []
        for header in self.metadata.sheet_headers:
            pre = header[:end_col]
            post = header[end_col:]
            new_header = pre + ['null'] + post
            new_sheet_headers.append(new_header)
        attribute_row = self.metadata.start_attribute_row
        new_sheet_headers[attribute_row][end_col] = attribute_type
        new_sheet_headers[attribute_row + 1][end_col] = attribute
        self.metadata.sheet_headers = new_sheet_headers
        new_data_row_list = []
        for data_row in self.data_row_list:
            pre = data_row[:end_col]
            post = data_row[end_col:]
            new_data_row = pre + [''] + post
            new_data_row_list.append(new_data_row)
        self.data_row_list = new_data_row_list
        self.data_row_list[attribute_row][end_col] = attribute_type
        self.data_row_list[attribute_row+1][end_col] = attribute
        for row in range(start_row, end_row):
            self.data_row_list[row][end_col] = value

def _verify_data_dir(dir_path):
    path = Path(dir_path)
    ok = path.exists() and path.is_dir()
    return ok


def main(project_name, experiment_name, output, download, deltas_path):
    builder = ExtractExperimentSpreadsheetWithChanges(output)
    ok = builder.set_up_project_experiment_metadata(project_name, experiment_name)
    if not ok:
        print("Invalid configuration of metadata or experiment/metadata mismatch. Quiting")
    else:
        if deltas_path:
            if builder.get_deltas_from_file(deltas_path):
                print("Applying deltas in the JSON file at " + deltas_path)
            else:
                print("Unexpectedly missing JSON file of delta list at " + deltas_path)
                print("Exiting.")
                exit(0)
        print("Writing experiment '" + builder.experiment.name
              + "' in project, '" + builder.project.name + ", to")
        print("spreadsheet at " + builder.output_path)
        builder.build_experiment_array()
        builder.write_spreadsheet()
        if download:
            print("Downloading process files to " + download)
            builder.download_process_files(download)


if __name__ == '__main__':
    time_stamp = '%s' % datetime.datetime.now()

    argv = sys.argv
    parser = argparse.ArgumentParser(
        description='Dump a project-experiment to a spreadsheet')
    parser.add_argument('proj', type=str, help="Project Name")
    parser.add_argument('exp', type=str, help="Experiment Name")
    parser.add_argument('output', type=str,
                        help='Path to output file')
    parser.add_argument('--deltas', type=str,
                        help="Path to file with json of delta-list; if none, no deltas are applied")
    parser.add_argument('--download', type=str,
                        help="Path to dir for downloading files; if none, files are not downloaded")
    args = parser.parse_args(argv[1:])

    args.output = os.path.abspath(args.output)

    if not args.output.endswith(".xlsx"):
        file = args.output + ".xlsx"

    if args.download:
        args.download = os.path.abspath(args.download)
        if not _verify_data_dir(args.download):
            print("Path for file download directory does not exist, ignoring: ", args.download)
            args.download = None

    if args.deltas:
        args.deltas = os.path.abspath(args.deltas)

    print("Output excel spreadsheet for experiment '" + args.exp + "' of project '" + args.proj + "'...")
    print("  to spreadsheet at " + args.output)
    if args.download:
        print("  with downloaded data going to " + args.download)
    if args.deltas:
        print("  applying the delta records in " + args.deltas)
    main(args.proj, args.exp, args.output, args.download, args.deltas)
