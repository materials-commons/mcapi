import openpyxl
import datetime
import os
from mcapi import create_project, get_all_templates

BASE_DATA_DIRECTORY = os.path.abspath("input_data/FromJake")


class WorkflowBuilderWE43FatigueLife:
    def __init__(self):
        self.test_sectioning = {}

    def parse_sheet(self, initiation_data_sheet, runout_data_sheet):
        initiation_data = self.parse_initiation_data(initiation_data_sheet)
        runout_data = self.parse_runout_data(runout_data_sheet)
        self.data = self.combine_data(initiation_data, runout_data)
        return self

    def build_workflow(self):
        print("build_workflow")
        self.template_table = self.make_template_table()

        self.create_project_experiment()

        self.load_files()

        we43 = self.create_we43_sample()
        self.section_we43 = self.create_we43_sectioning(we43)

        for sample in self.data:
            tests = self.data[sample]
            if len(tests) > 0:
                self.create_workflow(sample, tests)

        print(self.project.name)

    def create_project_experiment(self):
        time_stamp = '%s' % datetime.datetime.now()
        project_name = "WE43 Fatigue Life: " + time_stamp
        project_description = "WE43 Fatigue Life, created from script on " + time_stamp
        self.project = create_project(project_name, project_description)
        self.experiment = self.project.create_experiment("WE43 Fatigue Life", "Fatigue measurements on WE4s samples")

    def create_we43_sample(self):
        process_name = 'Obtain WE43 sample'
        template_id = self.template_id_with(self.template_table, 'Create Samples')
        process = self.experiment.create_process_from_template(template_id)
        process.rename(process_name)
        samples = process.create_samples(sample_names=['WE43'])
        return samples[0]

    def create_we43_sectioning(self, we43_sample):
        template_id = self.template_id_with(self.template_table, 'Sectioning')
        process = self.experiment.create_process_from_template(template_id)
        process.rename("Prepare WE43 Samples")
        process.add_input_samples_to_process([we43_sample])
        return process

    def create_workflow(self, sample_name, tests):
        print("create workflow for sample = " + sample_name)
        source_sample_name = tests[0]['material_condition']
        electropolish = (tests[0]['surface_condition'] == 'Electropolished')
        sample = self.get_sectioned_sample(source_sample_name, electropolish, sample_name)
        workflow_total_cycles = 0
        for test in tests:
            run_cycles = 0
            if test['test_type'] == "runout":
                run_cycles = test["test_parameters"]["cycle_count"]
            else:
                run_cycles = test["test_parameters"]["lifetime"]
            workflow_total_cycles += run_cycles
            test["test_parameters"]["total_cycles"] = workflow_total_cycles
            sample = self.create_fatigue_test(sample_name, sample, test)
            if test['test_type'] == "crack initiation":
                self.add_crack_initiation_data(sample_name, sample, test)

    def get_sectioned_sample(self, source_sample_name, electropolish, sample_name):
        if electropolish:
            source_sample_name += "-E"
        if not source_sample_name in self.test_sectioning:
            sample = self.add_source_sample(source_sample_name)
            sample = self.add_source_heat_treatment(source_sample_name, sample)
            if electropolish:
                sample = self.add_electropolishing(source_sample_name, sample)
            process = self.add_sectioning(source_sample_name, sample)
            self.test_sectioning[source_sample_name] = process
        process = self.test_sectioning[source_sample_name]
        section_samples = process.create_samples(sample_names=[sample_name])
        return section_samples[0]

    def add_source_sample(self, sample_name):
        process = self.section_we43
        samples = process.create_samples(sample_names=[sample_name])
        return samples[0]

    def add_source_heat_treatment(self, sample_name, sample):
        process_name = 'Heat Treat ' + sample_name
        template_id = self.template_id_with(self.template_table, 'Heat Treatment')
        process = self.experiment.create_process_from_template(template_id)
        process.rename(process_name)
        process.add_input_samples_to_process([sample])
        process = self.project.get_process_by_id(process.id)
        return process.output_samples[0]

    def add_electropolishing(self, sample_name, sample):
        process_name = 'Electropolish ' + sample_name
        template_id = self.template_id_with(self.template_table, 'Electropolishing')
        process = self.experiment.create_process_from_template(template_id)
        process.rename(process_name)
        process.add_input_samples_to_process([sample])
        process = self.project.get_process_by_id(process.id)
        return process.output_samples[0]

    def add_sectioning(self, source_sample_name, sample):
        template_id = self.template_id_with(self.template_table, 'Sectioning')
        process = self.experiment.create_process_from_template(template_id)
        process.rename("Section " + source_sample_name)
        process.add_input_samples_to_process([sample])
        return process

    def create_fatigue_test(self, sample_name, sample, data):
        process_name = 'Runout Fatigue Test ' + sample_name \
                       + "; Stress: " + str(data['test_parameters']['stress'])
        if data["test_type"] == "crack initiation":
            process_name = "Fatigue Life Test " + sample_name
        template_id = self.template_id_with(self.template_table, 'Ultrasonic Fatigue')
        process = self.experiment.create_process_from_template(template_id)
        process.rename(process_name)
        process.add_input_samples_to_process([sample])
        test_parameters = data['test_parameters']
        self.add_fatigue_data(process, test_parameters)
        process = self.project.get_process_by_id(process.id)
        return process.output_samples[0]

    def add_fatigue_data(self, process, data):
        annotation = "Control Unit = " + str(data['control_unit'])
        if 'cycle_count' in data:
            annotation += "; Run Cycle Count = " + '{:g}'.format(data['cycle_count'])
        if 'lifetime' in data:
            annotation += "; Run Lifetime = " + '{:g}'.format(data['lifetime'])
        annotation += "; Total Cycle Count = " + '{:g}'.format(data['total_cycles'])
        process = process.add_to_notes(annotation)

        stress = data['stress']
        environment = data['environment']
        process.set_value_of_setup_property('max_stress', stress)
        process.set_value_of_setup_property('test_environment', environment)
        process.update_setup_properties([
            'max_stress', 'test_environment'
        ])

    def add_crack_initiation_data(self,sample_name, sample, data):
        process_name = 'Crack Initiation ' + sample_name
        template_id = self.template_id_with(self.template_table, 'As Measured')
        process = self.experiment.create_process_from_template(template_id)
        process.rename(process_name)
        process.add_input_samples_to_process([sample])

        site_parameters = data['site_parameters']
        annotation = "Site: Initiation Type = " + site_parameters['type']
        if 'location_on_gage' in site_parameters and site_parameters['location_on_gage']:
            annotation += "; Location on Gage (mm from top) = " + site_parameters['location_on_gage']
        if 'location' in site_parameters and site_parameters['location']:
            annotation += "; Initiation Location = " + site_parameters['location']
        if 'fracture_surface_character' in site_parameters and site_parameters['fracture_surface_character']:
            annotation += "; Surface Character = " + site_parameters['fracture_surface_character']
        process.add_to_notes(annotation)

    def parse_initiation_data(self, data_sheet):
        list = []
        for col in data_sheet.iter_cols(min_col=2):
            values = []
            for cell in col:
                values.append(cell.value)
            if values[0]:
                list.append(self.make_initiation_data_entry(values))
        return list

    def parse_runout_data(self, data_sheet):
        list = []
        for col in data_sheet.iter_cols(min_col=2):
            values = []
            for cell in col:
                values.append(cell.value)
            if values[0]:
                list.append(self.make_runout_data_entry(values))
        return list

    def combine_data(self, initiation_data, runout_data):
        table = {}
        for element in runout_data:
            element['test_type'] = "runout"
            sample = element['sample']
            if not sample in table:
                table[sample] = []
            table[sample].append(element)
        for element in initiation_data:
            element['test_type'] = "crack initiation"
            sample = element['sample']
            if not sample in table:
                table[sample] = []
            table[sample].append(element)
        for sample in table:
            list = table[sample]
            list.sort(key=lambda e: e['testing_date'])
        return table

    def make_initiation_data_entry(self, values):
        return {
            "sample": values[0],
            "testing_date": values[1],
            "material_condition": values[2],
            "surface_condition": values[3],
            "test_parameters": {
                "stress": values[6],
                "lifetime": values[7],
                "environment": values[8],
                "step_sample": values[9],
                "control_unit": values[10],
            },
            "site_parameters": {
                "location_on_gage": values[13],
                "type": values[14],
                "location": values[15],
                "fracture_surface_character": values[16],
                "load_direction_inclination": values[17],
                "surface_inclination": values[18]
            },
            "grain_interior": {
                "schmid_factor": values[21],
                "dimension": values[22],
                "depth": values[23]
            },
            "grain_boundary": {
                "delta": values[26],
                "character": values[27]
            },
            "plotting": values[30],
            "notes": values[32]
        }

    def make_runout_data_entry(self, values):
        return {
            "sample": values[0],
            "testing_date": values[1],
            "material_condition": values[2],
            "surface_condition": values[3],
            "test_parameters": {
                "stress": values[6],
                "cycle_count": values[7],
                "environment": values[8],
                "step_sample": values[9],
                "control_unit": values[10],
            },
            "plotting": values[13],
            "notes": values[15]
        }

    def load_files(self):
        self.project.local_path = BASE_DATA_DIRECTORY
        path = os.path.join(BASE_DATA_DIRECTORY, "WE43FatigueLifeGraphs")
        self.project.add_directory_tree_by_local_path(path)

    def make_template_table(self):
        template_list = get_all_templates()
        table = {}
        for template in template_list:
            table[template.id] = template
        return table

    def template_id_with(self, table, match):
        found_id = None
        for key in table:
            if match in key:
                found_id = key
        return found_id


def main():
    path_and_name = 'input_data/FromJake/Fatigue Life/WE43 Fatigue Life.xlsx'
    wb = openpyxl.load_workbook(filename=path_and_name)
    ws1 = wb['Initiation Data']
    ws2 = wb['Runout Data']
    WorkflowBuilderWE43FatigueLife().parse_sheet(ws1, ws2).build_workflow()


if __name__ == '__main__':
    main()
